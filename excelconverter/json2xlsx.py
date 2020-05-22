# encoding: utf-8
import json, os
from openpyxl import load_workbook

def __parseArray(arr):
	ret = []
	for i in arr:
		if type(i) is dict:
			ret.append(__parseDict(i))
		elif type(i) is list:
			ret.append(__parseArray(i))
		else:
			ret.append(i)
	return ret

def __parseDict(d):
	ret = {}
	for k,v in d.items():
		if type(v) is dict:
			ret.update({ k+':'+kk : vv for kk,vv in __parseDict(v).items() })
		elif type(v) is list:
			tempret = { '%s[%d]'%(k,i) : item for i,item in enumerate(__parseArray(v)) }
			ret.update(__parseDict(tempret))
		else:
			ret[k] = v
	return ret

def convert(jsonfile):
	tsfilename = jsonfile.replace('.json', '.ts')
	spjsonname = jsonfile.replace('.json', '.sp')
	xlxsfilename = jsonfile.replace('.json', '.xlsx')
	sortedKeys = None
	sortedValues = None
	sheetName = 'Sheet'
	with open(jsonfile, 'r') as fjson:
		jsonObj = json.loads(fjson.read())
	with open(spjsonname, 'w') as spjson:
		if type(jsonObj) is dict:
			simpleJsonObj = __parseDict(jsonObj)
			sortedKeys = simpleJsonObj.keys()
			sortedKeys.sort()
			sortedValues = [ simpleJsonObj[k] for k in sortedKeys ]
			sheetName = 'SingleObject'
		elif type(jsonObj) is list:
			simpleJsonObj = __parseArray(jsonObj)
			if hasattr(simpleJsonObj[0], 'keys'):
				sortedKeys = simpleJsonObj[0].keys()
				sortedKeys.sort()
				sortedValues = [ simpleJsonObj[0][k] for k in sortedKeys ]
				sheetName = 'ObjectList'
			else:
				simpleJsonObj = [ { 'array' : o } for o in simpleJsonObj ]
				sortedKeys = [ 'array' ]
				sortedValues = [ simpleJsonObj[0]['array'] ]
				sheetName = 'PrimitiveList'
		else:
			raise Exception('json type must be dict or list')
		spjson.write(json.dumps(simpleJsonObj, indent=4))

	with open(tsfilename, 'w') as tsfile:
		tsfile.write('table {\n')
		for sk in sortedKeys:
			tsfile.write('	"%s";\n' % sk)
		tsfile.write('}')

	cmd = 'json2xlsx %s -j %s -o %s' % (tsfilename, spjsonname, xlxsfilename)
	if os.system(cmd) != 0:
		raise Exception('convert form json to xlxs error');
	print('%s 表成功生成' % xlxsfilename)

	# set the style of the excel sheet
	## freeze the first row
	wb = load_workbook(filename = xlxsfilename)
	ws = wb.active
	ws.title = sheetName
	i = 0
	for cell in ws[1]:
		typename = type(sortedValues[i]).__name__
		i = i + 1
		cell.value = "%s<%s>" % (cell.value, typename)
		cell.style = 'Headline 1'
	ws.freeze_panes = ws['A2']
	wb.save(xlxsfilename)

	# remove temp files
	os.remove(tsfilename)
	os.remove(spjsonname)
